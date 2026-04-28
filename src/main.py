from Bio import SeqIO
from Bio.SeqUtils.ProtParam import ProteinAnalysis
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import os

import matplotlib.pyplot as plt
image_files = []
kd_scale = {
    'A': 1.8, 'R': -4.5, 'N': -3.5, 'D': -3.5,
    'C': 2.5, 'Q': -3.5, 'E': -3.5, 'G': -0.4,
    'H': -3.2, 'I': 4.5, 'L': 3.8, 'K': -3.9,
    'M': 1.9, 'F': 2.8, 'P': -1.6, 'S': -0.8,
    'T': -0.7, 'W': -0.9, 'Y': -1.3, 'V': 4.2
}

def save_hydrophobicity_plot(sequence, filename, window_size=9):
    values = []
    positions = []

    for i in range(len(sequence) - window_size + 1):
        window = sequence[i:i+window_size]
        score = sum(kd_scale[aa] for aa in window) / window_size
        values.append(score)
        positions.append(i + window_size // 2)

    plt.figure()
    plt.plot(positions, values)
    plt.xlabel("Position")
    plt.ylabel("Hydrophobicity")
    plt.title(filename)
    plt.grid()

    plt.savefig(filename)
    plt.close()

input_fasta = "input.fasta"
output_excel = "output.xlsx"

wb = Workbook()
ws = wb.active

# Styles
bold = Font(bold=True)
right = Alignment(horizontal='right')
left = Alignment(horizontal='left')

row = 1  


for record in SeqIO.parse(input_fasta, "fasta"):
    start_row=row
    seq = str(record.seq)
    analysis = ProteinAnalysis(seq)

    # --- Protein Name ---
    ws.cell(row=row, column=1, value=record.id).font = bold
    row += 2

    # --- Composition ---
    ws.cell(row=row, column=1, value="Composition").font = bold
    row += 1

    aa_percent = analysis.amino_acids_percent

    for aa in sorted(aa_percent.keys()):
        ws.cell(row=row, column=1, value=aa).alignment = left

        cell = ws.cell(row=row, column=2, value=aa_percent[aa]/100)
        cell.number_format = '0.00%' 
        cell.alignment = right

        row += 1

    row += 1 

    # --- Parameters ---
    ws.cell(row=row, column=1, value="Parameter").font = bold
    ws.cell(row=row, column=2, value="Value").font = bold
    row += 1

    # Length
    ws.cell(row=row, column=1, value="Length").alignment = left
    ws.cell(row=row, column=2, value=len(seq)).alignment = right
    row += 1

    # Molecular Weight
    ws.cell(row=row, column=1, value="Molecular Weight").alignment = left
    cell = ws.cell(row=row, column=2, value=round(analysis.molecular_weight(), 2))
    cell.number_format = '0.00 "Da"'
    cell.alignment = right
    row += 1

    # Isoelectric Point
    ws.cell(row=row, column=1, value="Isoelectric Point").alignment = left
    ws.cell(row=row, column=2, value=round(analysis.isoelectric_point(), 2)).alignment = right
    row += 1

    # GRAVY
    ws.cell(row=row, column=1, value="GRAVY").alignment = left
    ws.cell(row=row, column=2, value=round(analysis.gravy(), 3)).alignment = right
    row += 1

    # Instability Index
    ws.cell(row=row, column=1, value="Instability Index").alignment = left
    ws.cell(row=row, column=2, value=round(analysis.instability_index(), 2)).alignment = right
    row += 1

    # Aromaticity
    ws.cell(row=row, column=1, value="Aromaticity").alignment = left
    ws.cell(row=row, column=2, value=round(analysis.aromaticity(), 3)).alignment = right
    row += 1

    # Secondary Structure
    helix, turn, sheet = analysis.secondary_structure_fraction()

    ws.cell(row=row, column=1, value="Helix").alignment = left
    cell = ws.cell(row=row, column=2, value=helix)
    cell.number_format = '0.00%'
    cell.alignment = right
    row += 1

    ws.cell(row=row, column=1, value="Turn").alignment = left
    cell = ws.cell(row=row, column=2, value=turn)
    cell.number_format = '0.00%'
    cell.alignment = right
    row += 1

    ws.cell(row=row, column=1, value="Sheet").alignment = left
    cell = ws.cell(row=row, column=2, value=sheet)
    cell.number_format = '0.00%'
    cell.alignment = right
    row += 1

    row += 2 

    # Save plot
    img_filename = f"{record.id}.png"
    save_hydrophobicity_plot(seq, record.id)

    # Insert image in Excel (right side)
    img = Image(img_filename)

    # Place image near current protein block
    img.anchor = f"D{start_row+5}"   # start_row = where protein starts
    ws.add_image(img)
    
    image_files.append(img_filename)


ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 20

wb.save(output_excel)

for file in image_files:
    os.remove(file)
print("Formatted analysis saved to output.xlsx")

import matplotlib.pyplot as plt


